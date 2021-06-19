import csv
import openpyxl as pyxl

wb = pyxl.load_workbook(
    "C:/Users/jamie/Downloads/Marketing Calendar 2011-2021.xlsx", "r"
)
sheet = wb["Filter List"]

row_count = sheet.max_row
column_count = sheet.max_column


def compare(x, y):
    if x[x.index(" ") :] == y[y.index(" ") :]:
        x_max = int(x.replace("*", "9")[: x.index(" ")])
        x_min = int(x.replace("*", "0")[: x.index(" ")])
        y_num = int(y[: y.index(" ")])
        if y_num >= x_min and y_num <= x_max:
            return True
    return False


with open("C:/Users/jamie/Downloads/August2021.csv", "r") as csv_file:

    csv_reader = csv.reader(csv_file)
    next(csv_reader)
    # Writes to new csv file with values omitted
    with open("C:/Users/jamie/Downloads/new_August2021.csv", "w") as new_file:
        csv_writer = csv.writer(new_file, delimiter=",", lineterminator="\n")
        # Writes to new file previous csv file but first checks if value should be added to csv
        with open(
            "C:/Users/jamie/Downloads/omit_new_August2021.csv", "w"
        ) as new_file_omit:
            csv_writer_omit = csv.writer(
                new_file_omit, delimiter=",", lineterminator="\n"
            )
            for line in csv_reader:
                check = True
                for row in range(1, row_count):
                    for column in range(1, column_count):
                        cell = str(sheet.cell(row=row, column=column).value).lower()
                        if "*" not in cell and cell in line[2].lower():
                            check = False
                            csv_writer_omit.writerow(line)
                        elif "*" in cell:
                            if compare(cell, line[2].lower()):
                                check = False
                                csv_writer_omit.writerow(line)
                if check == True:
                    csv_writer.writerow(line)