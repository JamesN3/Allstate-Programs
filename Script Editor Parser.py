import csv
import openpyxl as pyxl

wb = pyxl.load_workbook(
    "C:/Users/Public/Documents/Jamie's Work Folder/Filter street names.xlsx",
)
sheet = wb.worksheets[0]

row_count = sheet.max_row
column_count = sheet.max_column


def compare(x, y):
    if x[x.index(" ") :] == y[y.index(" ") :]:
        x_max = int(x.replace("*", "9")[: x.index(" ")])
        x_min = int(x.replace("*", "0")[: x.index(" ")])
        y_num = int(y[: y.index(" ")])
        if y_num >= x_min and y_num <= x_max:
            return True
    return False


with open(
    "C:/Users/Public/Documents/Jamie's Work Folder/July2021.csv", "r"
) as csv_file:

    csv_reader = csv.reader(csv_file)
    # Writes to new csv file with values omitted
    with open(
        "C:/Users/Public/Documents/Jamie's Work Folder/new_July2021.csv", "w"
    ) as new_file:
        # Writes to new file previous csv file but first checks if value should be added to csv
        csv_writer = csv.writer(new_file, delimiter=",", lineterminator="\n")
        for line in csv_reader:
            check = True
            for row in range(1, row_count):
                for column in range(1, column_count):
                    cell = str(sheet.cell(row=row, column=column).value).lower()
                    if "*" not in cell and cell in line[2].lower():
                        check = False
                        print(line)
                    elif "*" in cell:
                        if compare(cell, line[2].lower()):
                            check = False
                            print(line)
            if check == True:
                csv_writer.writerow(line)