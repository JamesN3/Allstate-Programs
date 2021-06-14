import csv
import openpyxl as pyxl

# Generate list of addresses in FIlter street names
wb = pyxl.load_workbook(
    "C:/Users/Public/Documents/Jamie's Work Folder/Filter street names.xlsx",
)

# Loads sheet from entire excel file
sheet = wb.worksheets[0]

# Finds bounds to search for values in compare_list
row_count = sheet.max_row
column_count = sheet.max_column

compare_list = []

# Functions made to create all possible permutations of asteriks
def compare_list_append(cell):
    count = cell.count("*")

    if count > 1:
        for i in range(0, 10):
            value1 = cell.replace("*", str(i), 1)
            compare_list_append(value1)
    elif count == 1:
        for i in range(0, 10):
            compare_list.append(cell.replace("*", str(i), 1))


# Loops through excel sheet to put all compare values into list
for row in range(1, row_count):
    for column in range(1, column_count):

        cell = str(sheet.cell(row=row, column=column).value).lower()

        if cell != "none":
            if "*" in cell:
                compare_list_append(cell)
            else:
                compare_list.append(cell)

# Opens previous csv file to compare
with open(
    "C:/Users/Public/Documents/Jamie's Work Folder/July2021.csv", "r"
) as csv_file:

    csv_reader = csv.reader(csv_file)
    # Writes to new csv file with values omitted
    with open(
        "C:/Users/Public/Documents/Jamie's Work Folder/new_July2021.csv", "w"
    ) as new_file:
        # Writes to new file previous csv file but first checks if value should be added to csv
        csv_writer = csv.writer(new_file, delimiter=",", lineterminator="\n")
        for line in csv_reader:
            check = True
            for compare_value in compare_list:
                if compare_value == line[2].lower():
                    check = False
            if check == True:
                csv_writer.writerow(line)
